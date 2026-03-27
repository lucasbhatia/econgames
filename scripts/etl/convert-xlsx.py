#!/usr/bin/env python3
"""
Convert large xlsx files to CSV using openpyxl's read-only mode.
This uses streaming (constant memory) so it handles 400K+ row files easily.

Run once before the Node.js ETL pipeline:
  python3 scripts/etl/convert-xlsx.py
"""

import csv
import os
from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'DATA')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'src', 'lib', 'data', 'processed')

os.makedirs(OUT_DIR, exist_ok=True)

FILES = [
    {
        'xlsx': 'GPS Races (December 24, 2025 – March 24, 2026).xlsx',
        'sheet': None,  # first sheet
        'csv': 'gps_races.csv',
    },
    {
        'xlsx': 'GPS PPs (relate to Starters PPs).xlsx',
        'sheet': None,
        'csv': 'gps_pps.csv',
    },
    {
        'xlsx': 'upcoming races.xlsx',
        'sheet': ' starters GPS PPs(if present) ',
        'csv': 'upcoming_gps_pps.csv',
    },
]

for file_info in FILES:
    xlsx_path = os.path.join(DATA_DIR, file_info['xlsx'])
    csv_path = os.path.join(OUT_DIR, file_info['csv'])

    print(f"\nConverting: {file_info['xlsx']}")
    print(f"  Sheet: {file_info['sheet'] or '(first)'}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[file_info['sheet']] if file_info['sheet'] else wb[wb.sheetnames[0]]

    rows_written = 0
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # Convert datetime objects to ISO strings
            cleaned = []
            for val in row:
                if hasattr(val, 'isoformat'):
                    cleaned.append(val.isoformat()[:10])
                elif val is None:
                    cleaned.append('')
                else:
                    cleaned.append(str(val).strip())
            writer.writerow(cleaned)
            rows_written += 1

    wb.close()

    size_mb = os.path.getsize(csv_path) / (1024 * 1024)
    print(f"  → {file_info['csv']}: {rows_written:,} rows ({size_mb:.1f} MB)")

print(f"\n✓ All conversions complete. CSV files in: {OUT_DIR}")
